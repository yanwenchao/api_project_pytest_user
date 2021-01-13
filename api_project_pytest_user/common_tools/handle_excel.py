"""
==================
Author: 严文超
2020/12/7 10:20
==================
"""
import openpyxl
import os
from common_tools.handle_path import DATA_DIR


class OperationExcel:

    def __init__(self, filename, sheetname):
        """
        :param filename: 要操作的excel文件名
        :param sheetname: 要操作的表单名
        """
        self.filename = filename
        self.sheetname = sheetname

    def open_excel(self):
        """打开操作表单"""
        self.workbook = openpyxl.load_workbook(filename=self.filename)
        self.sheet = self.workbook[self.sheetname]

    def read_excel(self):
        """读取excel表单数据"""
        self.open_excel()
        # 读取表单中所有数据
        rows = list(self.sheet.rows)
        # 获取第一行的表头数据
        case_title = [item.value for item in rows[0]]
        # 获取用例数据
        cases = []
        for row in rows[1:]:
            case_data = [item.value for item in row]
            case = dict(zip(case_title, case_data))
            cases.append(case)
        return cases

    def write_data(self, row, column, value,title=None):
        """
        :param row: 写入的行
        :param column: 写入的列
        :param value: 写入的内容
        :param title: 写入行的标题
        """
        self.open_excel()
        self.sheet.cell(row=1,column=column).value = title
        self.sheet.cell(row=row,column=column).value = value
        self.workbook.save(self.filename)

if __name__ == '__main__':
    wb = OperationExcel(os.path.join(DATA_DIR,"casedata.xlsx"),"Sheet1")
    cases = wb.read_excel()
    print(cases)